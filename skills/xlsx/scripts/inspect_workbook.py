from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_value(cell):
    value = cell.value
    if value is None:
        return ""
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def inspect_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    data_wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        data_ws = data_wb[ws.title]
        formulas = []
        errors = []
        preview = []
        for row in ws.iter_rows():
            preview_row = []
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.coordinate)
                    cached = data_ws[cell.coordinate].value
                    if isinstance(cached, str) and cached.startswith("#"):
                        errors.append({"cell": cell.coordinate, "error": cached})
                if len(preview) < 20 and cell.column <= 20:
                    preview_row.append(cell_value(cell))
            if len(preview) < 20 and any(item != "" for item in preview_row):
                preview.append(preview_row)
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
                "tables": list(ws.tables.keys()),
                "formula_count": len(formulas),
                "formula_cells_preview": formulas[:50],
                "cached_formula_errors": errors[:100],
                "preview_rows": preview,
            }
        )
    wb.close()
    data_wb.close()
    return {"workbook": path.name, "sheet_count": len(sheets), "sheets": sheets}


def csv_line(row: list[str]) -> str:
    return ",".join(f'"{item.replace(chr(34), chr(34) + chr(34))}"' if "," in item else item for item in row)


def to_markdown(payload: dict) -> str:
    lines = [f"# Workbook Inspection: {payload['workbook']}", ""]
    for sheet in payload["sheets"]:
        lines.extend(
            [
                f"## {sheet['name']}",
                "",
                f"- Size: {sheet['max_row']} rows x {sheet['max_column']} columns",
                f"- Formulas: {sheet['formula_count']}",
                f"- Tables: {', '.join(sheet['tables']) if sheet['tables'] else '(none)'}",
                f"- Merged ranges: {', '.join(sheet['merged_ranges']) if sheet['merged_ranges'] else '(none)'}",
                "",
                "### Preview",
                "",
                "```csv",
            ]
        )
        lines.extend(csv_line(row) for row in sheet["preview_rows"])
        lines.extend(["```", ""])
        if sheet["cached_formula_errors"]:
            lines.extend(["### Cached Formula Errors", ""])
            for error in sheet["cached_formula_errors"]:
                lines.append(f"- {error['cell']}: {error['error']}")
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python inspect_workbook.py <workbook.xlsx>", file=sys.stderr)
        return 2
    path = Path(sys.argv[1])
    if not path.is_file():
        print(f"Input workbook not found: {path}", file=sys.stderr)
        return 2
    payload = inspect_workbook(path)
    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    (out_dir / "inspection.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    (out_dir / "inspection.md").write_text(to_markdown(payload), encoding="utf-8")
    print(json.dumps({"status": "success", "outputs": ["outputs/inspection.json", "outputs/inspection.md"]}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
