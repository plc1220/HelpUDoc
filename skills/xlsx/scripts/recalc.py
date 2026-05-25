from __future__ import annotations

import json
import os
import platform
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
  Sub RecalculateAndSave()
    ThisComponent.calculateAll()
    ThisComponent.store()
    ThisComponent.close(True)
  End Sub
</script:module>"""


def soffice_env() -> dict:
    env = os.environ.copy()
    tmp_home = Path(env.get("HOME") or "tmp").resolve()
    tmp_home.mkdir(parents=True, exist_ok=True)
    env["HOME"] = str(tmp_home)
    env["SAL_USE_VCLPLUGIN"] = "svp"
    return env


def macro_dir() -> Path:
    if platform.system() == "Darwin":
        return Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    return Path.home() / ".config/libreoffice/4/user/basic/Standard"


def setup_macro() -> None:
    target_dir = macro_dir()
    target_file = target_dir / MACRO_FILENAME
    if target_file.exists() and "RecalculateAndSave" in target_file.read_text(encoding="utf-8", errors="replace"):
        return
    subprocess.run(["soffice", "--headless", "--terminate_after_init"], capture_output=True, timeout=15, env=soffice_env())
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file.write_text(RECALCULATE_MACRO, encoding="utf-8")


def scan_workbook(path: Path) -> dict:
    values = load_workbook(path, data_only=True)
    formulas = load_workbook(path, data_only=False)
    error_summary = {error: [] for error in EXCEL_ERRORS}
    formula_count = 0
    for sheet_name in formulas.sheetnames:
        formula_ws = formulas[sheet_name]
        value_ws = values[sheet_name]
        for row in formula_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                cached = value_ws[cell.coordinate].value
                if isinstance(cached, str):
                    for error in EXCEL_ERRORS:
                        if error in cached:
                            error_summary[error].append(f"{sheet_name}!{cell.coordinate}")
                            break
    values.close()
    formulas.close()
    compact_errors = {
        error: {"count": len(locations), "locations": locations[:50]}
        for error, locations in error_summary.items()
        if locations
    }
    total_errors = sum(item["count"] for item in compact_errors.values())
    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": compact_errors,
    }


def recalc(path: Path, timeout: int) -> dict:
    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    output_path = out_dir / "recalculated.xlsx"
    shutil.copy2(path, output_path)
    output_path.chmod(0o644)
    setup_macro()
    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(output_path.resolve()),
    ]
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    result = subprocess.run(cmd, capture_output=True, text=True, env=soffice_env())
    if result.returncode not in (0, 124):
        return {"error": result.stderr or "Unknown LibreOffice recalculation error"}
    payload = scan_workbook(output_path)
    payload["output"] = "outputs/recalculated.xlsx"
    return payload


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <workbook.xlsx> [timeout_seconds]", file=sys.stderr)
        return 2
    path = Path(sys.argv[1])
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    if not path.is_file():
        print(f"Input workbook not found: {path}", file=sys.stderr)
        return 2
    payload = recalc(path, timeout)
    Path("outputs").mkdir(exist_ok=True)
    (Path("outputs") / "recalc.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload))
    return 1 if "error" in payload else 0


if __name__ == "__main__":
    raise SystemExit(main())
