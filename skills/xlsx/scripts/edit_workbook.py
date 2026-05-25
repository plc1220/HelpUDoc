from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


def load_operations(argv: list[str]) -> tuple[Path | None, list[dict[str, Any]]]:
    input_path: Path | None = None
    operations: list[dict[str, Any]] | None = None
    idx = 1
    while idx < len(argv):
        key = argv[idx]
        value = argv[idx + 1] if idx + 1 < len(argv) else ""
        if key == "--input":
            input_path = Path(value)
            idx += 2
        elif key == "--operations":
            operations = json.loads(Path(value).read_text(encoding="utf-8"))
            idx += 2
        elif key == "--operations-json":
            operations = json.loads(value)
            idx += 2
        else:
            raise ValueError(f"Unknown argument: {key}")
    if operations is None:
        raise ValueError("Provide --operations <json-file> or --operations-json <json>")
    if not isinstance(operations, list):
        raise ValueError("Operations payload must be a JSON array")
    return input_path, operations


def get_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def apply_style(cell, op: dict[str, Any]) -> None:
    font = op.get("font")
    if isinstance(font, dict):
        cell.font = Font(
            name=font.get("name"),
            bold=font.get("bold"),
            italic=font.get("italic"),
            color=font.get("color"),
        )
    fill = op.get("fill")
    if isinstance(fill, str) and fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    alignment = op.get("alignment")
    if isinstance(alignment, dict):
        cell.alignment = Alignment(
            horizontal=alignment.get("horizontal"),
            vertical=alignment.get("vertical"),
            wrap_text=alignment.get("wrap_text"),
        )
    if op.get("number_format"):
        cell.number_format = str(op["number_format"])
    if op.get("comment"):
        cell.comment = Comment(str(op["comment"]), str(op.get("author") or "HelpUDoc"))


def set_cell(ws, op: dict[str, Any]) -> None:
    cell = ws[str(op["cell"])]
    if "formula" in op:
        formula = str(op["formula"])
        cell.value = formula if formula.startswith("=") else f"={formula}"
    elif "value" in op:
        cell.value = op["value"]
    apply_style(cell, op)


def apply_operation(wb, op: dict[str, Any]) -> None:
    kind = str(op.get("op") or "").strip()
    sheet_name = str(op.get("sheet") or "Sheet1")
    if kind == "add_sheet":
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        return
    ws = get_sheet(wb, sheet_name)
    if kind == "set_cell":
        set_cell(ws, op)
    elif kind == "append_row":
        ws.append(op.get("values") or [])
    elif kind == "insert_rows":
        ws.insert_rows(int(op["idx"]), int(op.get("amount") or 1))
    elif kind == "delete_rows":
        ws.delete_rows(int(op["idx"]), int(op.get("amount") or 1))
    elif kind == "insert_cols":
        ws.insert_cols(int(op["idx"]), int(op.get("amount") or 1))
    elif kind == "delete_cols":
        ws.delete_cols(int(op["idx"]), int(op.get("amount") or 1))
    elif kind == "set_column_width":
        ws.column_dimensions[str(op["column"])].width = float(op["width"])
    elif kind == "freeze_panes":
        ws.freeze_panes = str(op.get("cell") or "A2")
    else:
        raise ValueError(f"Unsupported operation: {kind}")


def main() -> int:
    try:
        input_path, operations = load_operations(sys.argv)
        if input_path is not None:
            if not input_path.is_file():
                raise ValueError(f"Input workbook not found: {input_path}")
            wb = load_workbook(input_path)
        else:
            wb = Workbook()
        for op in operations:
            if not isinstance(op, dict):
                raise ValueError("Each operation must be an object")
            apply_operation(wb, op)
        out_dir = Path("outputs")
        out_dir.mkdir(exist_ok=True)
        output = out_dir / "workbook.xlsx"
        wb.save(output)
        payload = {"status": "success", "operation_count": len(operations), "output": "outputs/workbook.xlsx"}
    except Exception as exc:
        payload = {"status": "error", "error": str(exc)}
    Path("outputs").mkdir(exist_ok=True)
    (Path("outputs") / "result.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload))
    return 0 if payload["status"] == "success" else 1


if __name__ == "__main__":
    raise SystemExit(main())
